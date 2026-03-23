#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Импорт XLSX (выгрузка из dgdat) → SQLite.

Читает все .xlsx из XLSX_FOLDER, создаёт нормализованную БД в DB_PATH.
Идемпотентен: повторный запуск не создаёт дублей.
"""

import hashlib
import os
import re
import sqlite3
import sys
import time
from urllib.parse import urlparse

from openpyxl import load_workbook

# ============================================================
# НАСТРОЙКИ
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

XLSX_FOLDER = os.path.join(SCRIPT_DIR, "output")
DB_PATH = os.path.join(SCRIPT_DIR, "data", "local.db")

# Индексы колонок (0-based) в XLSX
COL = {
    "id": 0,
    "name": 1,
    "city": 2,
    "section": 3,
    "subsection": 4,
    "rubric": 5,
    "phones": 6,
    "faxes": 7,
    "email": 8,
    "website": 9,
    "address": 10,
    "postal_code": 11,
    "payment_types": 12,
    "working_hours": 13,
    "building_name": 14,
    "building_type": 15,
    "vk": 16,
    "facebook": 17,
    "skype": 18,
    "twitter": 19,
    "instagram": 20,
    "icq": 21,
    "jabber": 22,
}

SOCIAL_COLS = [
    ("vk", 16),
    ("facebook", 17),
    ("skype", 18),
    ("twitter", 19),
    ("instagram", 20),
    ("icq", 21),
    ("jabber", 22),
]

# ============================================================
# SCHEMA
# ============================================================

SCHEMA = """
CREATE TABLE IF NOT EXISTS companies (
    id INTEGER PRIMARY KEY,
    name TEXT NOT NULL,
    city TEXT,
    website TEXT,
    domain TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS company_aliases (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    company_id INTEGER NOT NULL REFERENCES companies(id),
    name TEXT NOT NULL,
    UNIQUE(company_id, name)
);

CREATE TABLE IF NOT EXISTS branches (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    company_id INTEGER NOT NULL REFERENCES companies(id),
    address TEXT,
    postal_code TEXT,
    working_hours TEXT,
    building_name TEXT,
    building_type TEXT,
    branch_hash TEXT UNIQUE
);

CREATE TABLE IF NOT EXISTS phones (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    branch_id INTEGER NOT NULL REFERENCES branches(id),
    phone TEXT NOT NULL,
    UNIQUE(branch_id, phone)
);

CREATE TABLE IF NOT EXISTS emails (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    company_id INTEGER NOT NULL REFERENCES companies(id),
    email TEXT NOT NULL,
    UNIQUE(company_id, email)
);

CREATE TABLE IF NOT EXISTS socials (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    company_id INTEGER NOT NULL REFERENCES companies(id),
    type TEXT NOT NULL,
    url TEXT NOT NULL,
    UNIQUE(company_id, type, url)
);

CREATE TABLE IF NOT EXISTS categories (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    parent_id INTEGER REFERENCES categories(id),
    UNIQUE(name, parent_id)
);

CREATE TABLE IF NOT EXISTS company_categories (
    company_id INTEGER NOT NULL REFERENCES companies(id),
    category_id INTEGER NOT NULL REFERENCES categories(id),
    PRIMARY KEY(company_id, category_id)
);
"""

# ============================================================
# УТИЛИТЫ
# ============================================================


def is_valid(value) -> bool:
    """Проверяет, является ли значение валидным (не None и не пустой текст)."""
    if value is None:
        return False
    value_str = str(value).strip()
    return value_str != ""


def clean_text(value: str) -> str | None:
    """Очищает текст: удаляет переводы строк, нормализует пробелы."""
    if not value:
        return None
    text = str(value).replace("\n", " ").strip()
    # Нормализуем множественные пробелы
    text = re.sub(r'\s+', ' ', text)
    return text if text else None


def extract_domain(website: str) -> str | None:
    """Извлекает домен из URL: убирает http(s), www, приводит к lowercase.

    Сначала очищает текст от переводов строк, затем извлекает домен с помощью regex.
    """
    if not website:
        return None

    # Очищаем текст от переводов строк
    text = str(website).replace("\n", " ").lower().strip()
    if not text:
        return None

    # Пытаемся найти домен используя regex: domain.tld
    match = re.search(r'(?:https?://)?(?:www\.)?([a-z0-9.-]+\.[a-z]{2,})', text)
    if match:
        domain = match.group(1)
        # Убираем www. если осталось
        if domain.startswith("www."):
            domain = domain[4:]
        return domain if domain else None

    return None


def split_values(text: str) -> list[str]:
    """Разделяет строку по , ; \\n и возвращает очищенные непустые значения."""
    if not text:
        return []
    parts = re.split(r'[,;\n]+', str(text))
    return [p.strip() for p in parts if p.strip()]


def normalize_address(addr: str) -> str:
    """Нормализует адрес для хеширования: lowercase, без лишних пробелов."""
    if not addr:
        return ""
    return re.sub(r'\s+', ' ', str(addr).strip().lower())


def make_branch_hash(company_id: int, address: str) -> str:
    """MD5-хеш ключа филиала: company_id + нормализованный адрес."""
    key = f"{company_id}_{normalize_address(address)}"
    return hashlib.md5(key.encode("utf-8")).hexdigest()


def cell_str(val) -> str:
    """Безопасное преобразование ячейки в строку."""
    if val is None:
        return ""
    return str(val).strip()


# ============================================================
# ИНИЦИАЛИЗАЦИЯ БД
# ============================================================


def init_db(db_path: str) -> sqlite3.Connection:
    """Создаёт БД и таблицы (если не существуют). Возвращает соединение."""
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.executescript(SCHEMA)
    return conn


# ============================================================
# ОБРАБОТКА КАТЕГОРИЙ
# ============================================================


def get_or_create_category(cur: sqlite3.Cursor, name: str, parent_id: int | None) -> int:
    """Возвращает id категории, создавая при необходимости."""
    if parent_id is None:
        cur.execute(
            "SELECT id FROM categories WHERE name = ? AND parent_id IS NULL", (name,)
        )
    else:
        cur.execute(
            "SELECT id FROM categories WHERE name = ? AND parent_id = ?",
            (name, parent_id),
        )
    row = cur.fetchone()
    if row:
        return row[0]

    cur.execute(
        "INSERT INTO categories (name, parent_id) VALUES (?, ?)", (name, parent_id)
    )
    return cur.lastrowid


def process_categories(cur: sqlite3.Cursor, company_id: int, section: str, subsection: str, rubric: str):
    """Разбирает иерархию категорий и связывает с компанией."""
    leaf_ids = set()

    # Каждое из трёх полей может содержать несколько значений через \n
    sections = split_values(section) if section else []
    subsections = split_values(subsection) if subsection else []
    rubrics = split_values(rubric) if rubric else []

    # Строим дерево: Раздел → Подраздел → Рубрика
    # Каждое значение может содержать "/" для вложенности
    def resolve_chain(names: list[str], parent_id: int | None) -> list[int]:
        """Разбивает каждое имя по '/' и создаёт цепочку."""
        result = []
        for raw in names:
            parts = [p.strip() for p in raw.split("/") if p.strip()]
            pid = parent_id
            cat_id = pid
            for part in parts:
                cat_id = get_or_create_category(cur, part, pid)
                pid = cat_id
            if cat_id is not None:
                result.append(cat_id)
        return result

    section_ids = resolve_chain(sections, None)

    if subsections:
        sub_ids = []
        parents = section_ids if section_ids else [None]
        for parent in parents:
            sub_ids.extend(resolve_chain(subsections, parent))
        if rubrics:
            rub_parents = sub_ids if sub_ids else parents
            for parent in rub_parents:
                leaf_ids.update(resolve_chain(rubrics, parent))
        else:
            leaf_ids.update(sub_ids)
    elif rubrics:
        parents = section_ids if section_ids else [None]
        for parent in parents:
            leaf_ids.update(resolve_chain(rubrics, parent))
    else:
        leaf_ids.update(section_ids)

    for cat_id in leaf_ids:
        cur.execute(
            "INSERT INTO company_categories (company_id, category_id) VALUES (?, ?) ON CONFLICT(company_id, category_id) DO NOTHING",
            (company_id, cat_id),
        )


# ============================================================
# ОБРАБОТКА СТРОКИ
# ============================================================


def process_row(cur: sqlite3.Cursor, values: list, stats: dict):
    """Обрабатывает одну строку XLSX.

    Логика:
    1. Вставляем компанию если её нет (INSERT ON CONFLICT DO NOTHING)
    2. Обновляем только валидные значения (не NULL, не пустые)
    3. Используем CASE WHEN для условного обновления
    """
    company_id = values[COL["id"]]
    if not company_id:
        return

    company_id = int(company_id)
    name = cell_str(values[COL["name"]])
    city = cell_str(values[COL["city"]]) or None
    website_raw = cell_str(values[COL["website"]]) or None

    # Очищаем website и извлекаем домен
    website = clean_text(website_raw) if website_raw else None
    domain = extract_domain(website) if website else None

    # --- Company: INSERT ON CONFLICT DO NOTHING ---
    # Сначала проверяем, существует ли уже компания
    cur.execute("SELECT 1 FROM companies WHERE id = ?", (company_id,))
    is_existing = cur.fetchone() is not None

    if not is_existing:
        # Вставляем новую компанию
        cur.execute(
            """INSERT INTO companies (id, name, city, website, domain, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)""",
            (company_id, name, city, website, domain),
        )
        stats["companies_new"] += 1
    else:
        # Компания уже существовала, проверяем если изменилось имя
        cur.execute("SELECT name FROM companies WHERE id = ?", (company_id,))
        row = cur.fetchone()
        existing_name = row[0] if row else None

        if name and existing_name and name != existing_name:
            cur.execute(
                "INSERT INTO company_aliases (company_id, name) VALUES (?, ?) ON CONFLICT(company_id, name) DO NOTHING",
                (company_id, name),
            )

    # --- Company: UPDATE с условиями (только валидные значения) ---
    # Используем CASE WHEN чтобы обновить только если новое значение валидно
    cur.execute(
        """UPDATE companies
           SET name = CASE WHEN ? IS NOT NULL AND ? != '' THEN ? ELSE name END,
               city = CASE WHEN ? IS NOT NULL AND ? != '' THEN ? ELSE city END,
               website = CASE WHEN ? IS NOT NULL AND ? != '' THEN ? ELSE website END,
               domain = CASE WHEN ? IS NOT NULL AND ? != '' THEN ? ELSE domain END,
               updated_at = CURRENT_TIMESTAMP
           WHERE id = ?""",
        (
            name, name, name,
            city, city, city,
            website, website, website,
            domain, domain, domain,
            company_id
        ),
    )

    # --- Branch ---
    address = cell_str(values[COL["address"]]) or None
    postal_code = cell_str(values[COL["postal_code"]]) or None
    working_hours = cell_str(values[COL["working_hours"]]) or None
    building_name = cell_str(values[COL["building_name"]]) or None
    building_type = cell_str(values[COL["building_type"]]) or None
    branch_hash = make_branch_hash(company_id, address or "")

    cur.execute("SELECT id FROM branches WHERE branch_hash = ?", (branch_hash,))
    branch_row = cur.fetchone()

    if branch_row:
        branch_id = branch_row[0]
        # UPDATE ветки с условиями для валидности
        cur.execute(
            """UPDATE branches
               SET address = CASE WHEN ? IS NOT NULL AND ? != '' THEN ? ELSE address END,
                   postal_code = CASE WHEN ? IS NOT NULL AND ? != '' THEN ? ELSE postal_code END,
                   working_hours = CASE WHEN ? IS NOT NULL AND ? != '' THEN ? ELSE working_hours END,
                   building_name = CASE WHEN ? IS NOT NULL AND ? != '' THEN ? ELSE building_name END,
                   building_type = CASE WHEN ? IS NOT NULL AND ? != '' THEN ? ELSE building_type END
               WHERE id = ?""",
            (
                address, address, address,
                postal_code, postal_code, postal_code,
                working_hours, working_hours, working_hours,
                building_name, building_name, building_name,
                building_type, building_type, building_type,
                branch_id
            ),
        )
    else:
        # Новая ветка
        cur.execute(
            """INSERT INTO branches (company_id, address, postal_code, working_hours,
                                     building_name, building_type, branch_hash)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (company_id, address, postal_code, working_hours, building_name, building_type, branch_hash),
        )
        branch_id = cur.lastrowid
        stats["branches_new"] += 1

    # --- Phones + Faxes → phones ---
    for col_key in ("phones", "faxes"):
        for phone in split_values(cell_str(values[COL[col_key]])):
            cur.execute(
                "INSERT INTO phones (branch_id, phone) VALUES (?, ?) ON CONFLICT(branch_id, phone) DO NOTHING",
                (branch_id, phone),
            )

    # --- Emails ---
    for email in split_values(cell_str(values[COL["email"]])):
        email_lower = email.lower()
        cur.execute(
            "INSERT INTO emails (company_id, email) VALUES (?, ?) ON CONFLICT(company_id, email) DO NOTHING",
            (company_id, email_lower),
        )

    # --- Socials ---
    for social_type, col_idx in SOCIAL_COLS:
        url = cell_str(values[col_idx])
        if is_valid(url):
            cur.execute(
                "INSERT INTO socials (company_id, type, url) VALUES (?, ?, ?) ON CONFLICT(company_id, type, url) DO NOTHING",
                (company_id, social_type, url),
            )

    # --- Categories ---
    section = cell_str(values[COL["section"]])
    subsection = cell_str(values[COL["subsection"]])
    rubric = cell_str(values[COL["rubric"]])
    if section or subsection or rubric:
        process_categories(cur, company_id, section, subsection, rubric)


# ============================================================
# ОБРАБОТКА ФАЙЛА
# ============================================================


def process_file(conn: sqlite3.Connection, xlsx_path: str) -> dict:
    """Обрабатывает один XLSX-файл. Возвращает статистику."""
    stats = {"rows": 0, "companies_new": 0, "branches_new": 0}
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    cur = conn.cursor()
    num_cols = len(COL)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=num_cols, values_only=True), 1):
        values = list(row)
        if not values or not values[0]:
            continue
        process_row(cur, values, stats)
        stats["rows"] += 1

        if row_idx % 1000 == 0:
            print(f"    {row_idx} строк...")

    conn.commit()
    wb.close()
    return stats


# ============================================================
# MAIN
# ============================================================


def main():
    if not os.path.isdir(XLSX_FOLDER):
        print(f"Папка не найдена: {XLSX_FOLDER}")
        sys.exit(1)

    xlsx_files = sorted(
        f for f in os.listdir(XLSX_FOLDER) if f.lower().endswith(".xlsx")
    )
    if not xlsx_files:
        print(f"XLSX-файлов не найдено в {XLSX_FOLDER}")
        sys.exit(1)

    print(f"Найдено файлов: {len(xlsx_files)}")
    print(f"База данных: {DB_PATH}")
    print()

    conn = init_db(DB_PATH)
    total_stats = {"files": 0, "rows": 0, "companies_new": 0, "branches_new": 0}
    t0 = time.time()

    for fname in xlsx_files:
        path = os.path.join(XLSX_FOLDER, fname)
        print(f"  Импорт: {fname}")
        stats = process_file(conn, path)
        total_stats["files"] += 1
        total_stats["rows"] += stats["rows"]
        total_stats["companies_new"] += stats["companies_new"]
        total_stats["branches_new"] += stats["branches_new"]
        print(f"    Строк: {stats['rows']}, новых компаний: {stats['companies_new']}, новых филиалов: {stats['branches_new']}")

    conn.close()
    elapsed = time.time() - t0

    print()
    print("=" * 50)
    print(f"Итого:")
    print(f"  Файлов: {total_stats['files']}")
    print(f"  Строк обработано: {total_stats['rows']}")
    print(f"  Новых компаний: {total_stats['companies_new']}")
    print(f"  Новых филиалов: {total_stats['branches_new']}")
    print(f"  Время: {elapsed:.1f} сек")
    print(f"  БД: {DB_PATH}")


if __name__ == "__main__":
    main()
