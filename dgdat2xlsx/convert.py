#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Парсер файлов .dgdat (2ГИС) → XLSX

Порт PHP-проекта https://github.com/mbry/DgdatToXlsx/
При использовании алгоритмов или части кода ссылка на первоисточник обязательна!
"""

import struct
import sys
import os
import json
import xml.etree.ElementTree as ET
from io import BytesIO

import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection

# ============================================================
# НАСТРОЙКИ ПУТЕЙ — измените при необходимости
# ============================================================

# Папка с входными .dgdat файлами
# DGDAT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'download')
DGDAT_DIR = 'C:\\Program Files (x86)\\2gis\\3.0'

# Папка для выходных .xlsx файлов
# XLSX_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
XLSX_DIR = 'D:\\2GIS'

# Паттерн для удаления недопустимых символов XML (управляющие символы кроме \t, \n, \r)
ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')


# ============================================================
# Низкоуровневое чтение бинарных данных
# ============================================================

class BinaryReader:
    """Чтение из файла (аналог PHP fread / fseek / ftell)."""

    def __init__(self, fp):
        self.fp = fp

    def read_long(self):
        v = self.fp.read(4)
        if len(v) < 4:
            raise EOFError("Неожиданный конец файла при чтении Long")
        return struct.unpack('<L', v)[0]

    def read_byte(self):
        v = self.fp.read(1)
        if len(v) < 1:
            raise EOFError("Неожиданный конец файла при чтении Byte")
        return struct.unpack('B', v)[0]

    def read_string(self, length):
        if length <= 0:
            return b''
        return self.fp.read(length)

    def read_packed_value(self):
        size = self.read_byte()
        if size >= 0xF0:
            b1 = self.read_byte()
            b2 = self.read_byte()
            b3 = self.read_byte()
            b4 = self.read_byte()
            return (b1 << 24) | (b2 << 16) | (b3 << 8) | b4
        if size >= 0xE0:
            b2 = self.read_byte()
            b3 = self.read_byte()
            b4 = self.read_byte()
            return ((size ^ 0xE0) << 24) | (b2 << 16) | (b3 << 8) | b4
        if size >= 0xC0:
            b2 = self.read_byte()
            b3 = self.read_byte()
            return ((size ^ 0xC0) << 16) | (b2 << 8) | b3
        if size & 0x80:
            b2 = self.read_byte()
            return ((size ^ 0x80) << 8) | b2
        return size

    def tell(self):
        return self.fp.tell()

    def seek(self, pos):
        self.fp.seek(pos)


def get_packed_value(data: bytes, offset: int):
    """Читает packed value из буфера, возвращает (значение, новый_offset)."""
    if offset >= len(data):
        return 0, offset
    size = data[offset]
    offset += 1
    if size >= 0xF0:
        if offset + 4 > len(data):
            return size, offset
        b1 = data[offset]; b2 = data[offset+1]; b3 = data[offset+2]; b4 = data[offset+3]
        offset += 4
        return (b1 << 24) | (b2 << 16) | (b3 << 8) | b4, offset
    if size >= 0xE0:
        if offset + 3 > len(data):
            return size, offset
        b2 = data[offset]; b3 = data[offset+1]; b4 = data[offset+2]
        offset += 3
        return ((size ^ 0xE0) << 24) | (b2 << 16) | (b3 << 8) | b4, offset
    if size >= 0xC0:
        if offset + 2 > len(data):
            return size, offset
        b2 = data[offset]; b3 = data[offset+1]
        offset += 2
        return ((size ^ 0xC0) << 16) | (b2 << 8) | b3, offset
    if size & 0x80:
        if offset + 1 > len(data):
            return size, offset
        b2 = data[offset]
        offset += 1
        return ((size ^ 0x80) << 8) | b2, offset
    return size, offset


# ============================================================
# Распаковка строк (UnpackWideString)
# ============================================================

def unpack_wide_string(data: bytes) -> bytes:
    """Распаковка строки в UTF-16LE из сжатого формата dgdat."""
    if len(data) == 0:
        return b''
    offset = 0
    x1, offset = get_packed_value(data, offset)
    x2, offset = get_packed_value(data, offset)

    z = bytearray()
    for i in range(x2):
        if offset < len(data):
            z.append(data[offset])
            z.append(0)
            offset += 1

    if offset < len(data):
        mcount = data[offset]
        offset += 1

        arr = []
        for i in range(mcount):
            if offset < len(data):
                arr.append(data[offset])
                offset += 1

        if mcount > 0:
            ziter = 0
            while offset < len(data):
                v = data[offset]
                offset += 1
                count = v // mcount
                off_mod = v % mcount

                if count == 0:
                    l = len(z)
                    for i in range(ziter, l, 2):
                        if i + 1 < len(z):
                            z[i + 1] = arr[off_mod]
                else:
                    for i in range(count):
                        if ziter + 1 < len(z):
                            z[ziter + 1] = arr[off_mod]
                        ziter += 2

    return bytes(z)


# ============================================================
# Обработка таблиц
# ============================================================

DO_NOT_EXPORT = [
    "route_", "ctr_", "geo_", "chm_store", "org_banner",
    "back_splash", "banner_", "road_", "interchange_", "logo_picture", "pk_"
]


def process_table(name: str, data: bytes, datadir: dict):
    """Разбирает таблицу данных (аналог ProcessTable в PHP)."""
    for prefix in DO_NOT_EXPORT:
        if name.startswith(prefix):
            return

    offset = 0
    tbllen, offset = get_packed_value(data, offset)
    tbl = data[offset:offset + tbllen]
    rest = data[offset + tbllen:]

    tbl_offset = 0
    while tbl_offset < len(tbl):
        chunk_len = tbl[tbl_offset]
        tbl_offset += 1
        chunk_name = tbl[tbl_offset:tbl_offset + chunk_len].decode('ascii', errors='replace')
        tbl_offset += chunk_len

        if chunk_name == "data":
            if name not in datadir:
                datadir[name] = {}
            datadir[name]["data"] = data
            return

        size, new_tbl_offset = get_packed_value(tbl, tbl_offset)
        tbl_offset = new_tbl_offset

        p = rest[:size]
        rest = rest[size:]

        dexor_table(name, chunk_name, p, datadir, need_decode=0)


def dexor_table(name: str, fieldname: str, data: bytes, datadir: dict, need_decode=1) -> bytes:
    """Дешифровка и сохранение поля таблицы (аналог DexorTable)."""
    if fieldname != "":
        if name not in datadir:
            datadir[name] = {}
        datadir[name][fieldname] = data

    if len(data) == 0:
        return b''

    offset = 0
    tbllen, offset = get_packed_value(data, offset)
    if tbllen == 0:
        return b''

    skip_len, offset = get_packed_value(data, offset)
    offset += skip_len
    dat_len, offset = get_packed_value(data, offset)

    start = tbllen + 1
    copy = bytearray(data)
    dat = bytearray()

    for i in range(start, start + dat_len):
        if i >= len(copy):
            break
        ch = copy[i]
        if need_decode == 1:
            decoded = ch ^ 0xC5
            dat.append(decoded)
            copy[i] = decoded
        else:
            dat.append(ch)

    return bytes(dat)


# ============================================================
# Экспорт полей (ExportField)
# ============================================================

def export_field(name: str, field: str, datadir: dict,
                 need_decode=1, pair_decode=0) -> dict:
    """Извлекает поле из распарсенных данных (аналог ExportField)."""
    if name not in datadir or field not in datadir[name]:
        return {}

    dat_raw = datadir[name][field]
    dat = dexor_table(name, field, dat_raw, datadir, need_decode)

    if len(dat) == 0:
        return {}

    afield = {}

    if pair_decode == 10:
        i = 1
        for k in range(0, len(dat), 4):
            temp = dat[k:k + 4]
            if len(temp) < 4:
                break
            afield[i] = struct.unpack('<L', temp)[0]
            i += 1
        return afield

    if pair_decode == 1:
        i = 1
        k = 0
        while k < len(dat):
            repeat, new_k = get_packed_value(dat, k)
            value, new_k = get_packed_value(dat, new_k)
            for _ in range(repeat):
                afield[i] = value
                i += 1
            k = new_k
        return afield

    if pair_decode == 4:
        # PHP: for($k=0; ...; $k++) { repeat=GPV; nm=consumed;
        #   len=GPV; xlen=GPV; afield[repeat]=xlen; $k = $k + $len + $nm; }
        # +$k++ = +1
        i = 1
        k = 0
        while k < len(dat):
            init_k = k
            repeat, k = get_packed_value(dat, k)
            nm = k - init_k
            length, k = get_packed_value(dat, k)
            xlen, k = get_packed_value(dat, k)
            afield[repeat] = xlen
            k = init_k + length + nm + 1
        return afield

    if pair_decode == 2:
        i = 1
        k = 0
        while k < len(dat):
            value, k = get_packed_value(dat, k)
            afield[i] = value
            i += 1
        return afield

    if pair_decode == 3:
        # В PHP: for($k=0; ...; $k++) { repeat=GPV; nm=bytes_consumed;
        #   len=GPV; ... $k = $k + $len + $nm; }
        # Плюс $k++ из for = +1. Но в случае len>0x7f $len++ делает +1 к len.
        i = 1
        k = 0
        while k < len(dat):
            init_k = k
            repeat, k = get_packed_value(dat, k)
            nm = k - init_k  # байт на repeat

            len_start = k
            length, k = get_packed_value(dat, k)

            if length == 0:
                x_str = ""
            elif length > 0x7F:
                raw = dat[len_start:len_start + length + 2]
                x_bytes = unpack_wide_string(raw)
                x_str = x_bytes.decode('utf-16-le', errors='replace')
                length += 1  # как в PHP: $len++
            else:
                raw = dat[len_start:len_start + length + 1]
                x_bytes = unpack_wide_string(raw)
                x_str = x_bytes.decode('utf-16-le', errors='replace')

            for _ in range(repeat):
                afield[i] = x_str
                i += 1

            # PHP: $k = $k + $len + $nm; затем $k++ итого +1
            k = init_k + length + nm + 1
        return afield

    # pair_decode == 0 (default) — строки
    # В PHP: for($k=0; ...; $k++) { ... $k = $k + $len; }
    # Итого за итерацию: len==0 -> +1; len<=0x7f -> len+1; len>0x7f -> len+2
    i = 1
    k = 0
    while k < len(dat):
        length, after_prefix = get_packed_value(dat, k)
        prefix_size = after_prefix - k  # сколько байт занял packed value

        if length == 0:
            x_str = ""
            k += prefix_size
        elif length > 0x7F:
            raw = dat[k:k + length + 2]
            x_bytes = unpack_wide_string(raw)
            x_str = x_bytes.decode('utf-16-le', errors='replace')
            k += length + 2  # 2-byte prefix + length bytes data
        else:
            raw = dat[k:k + length + 1]
            x_bytes = unpack_wide_string(raw)
            x_str = x_bytes.decode('utf-16-le', errors='replace')
            k += length + 1  # 1-byte prefix + length bytes data

        afield[i] = x_str
        i += 1

    return afield


# ============================================================
# Основной парсер .dgdat файла
# ============================================================

def parse_dgdat(filepath: str) -> dict:
    """Парсит .dgdat файл и возвращает словарь dump с данными организаций."""

    datadir = {}
    prop = {}

    with open(filepath, 'rb') as f:
        reader = BinaryReader(f)

        # Проверка магического числа
        magic = reader.read_long()
        ef = reader.read_byte()

        if hex(magic) != '0x46444707' or ef != 239:
            raise ValueError("Это не файл данных 2ГИС (.dgdat)")

        reader.read_long()
        reader.read_long()

        reader.read_packed_value()
        reader.read_packed_value()
        reader.read_packed_value()
        reader.read_packed_value()

        # Первая таблица директории
        tbllen = reader.read_byte()
        tbl = reader.read_string(tbllen)

        startdir = []
        root = None
        optroot = None

        offset = 0
        while offset < len(tbl):
            chunk_len = tbl[offset]
            offset += 1
            chunk_name = tbl[offset:offset + chunk_len].decode('ascii', errors='replace')
            offset += chunk_len

            size, offset = get_packed_value(tbl, offset)

            startdir.append({
                'name': chunk_name,
                'size': size,
                'offset': reader.tell()
            })

            temp = reader.read_string(size)

            inset = ["name", "cpt", "fbn", "lang", "stat"]
            if chunk_name in inset:
                ws = unpack_wide_string(temp)
                prop[chunk_name] = ws.decode('utf-16-le', errors='replace')

        reader.read_packed_value()

        # Вторая таблица директории
        tbllen = reader.read_packed_value()
        tbl = reader.read_string(tbllen)

        offset = 0
        while offset < len(tbl):
            chunk_len = tbl[offset]
            offset += 1
            chunk_name = tbl[offset:offset + chunk_len].decode('ascii', errors='replace')
            offset += chunk_len

            size, offset = get_packed_value(tbl, offset)

            startdir.append({
                'name': chunk_name,
                'size': size,
                'offset': reader.tell()
            })

            if chunk_name == "data":
                root = reader.tell()
            elif chunk_name == "opt":
                optroot = reader.tell()

            reader.read_string(size)

        # Обработка корневой таблицы (data)
        if root is None:
            raise ValueError("Таблица 'data' не найдена в файле")

        reader.seek(root)

        tbllen = reader.read_packed_value()
        tbl = reader.read_string(tbllen)

        offset = 0
        while offset < len(tbl):
            chunk_len = tbl[offset]
            offset += 1
            chunk_name = tbl[offset:offset + chunk_len].decode('ascii', errors='replace')
            offset += chunk_len

            size, offset = get_packed_value(tbl, offset)

            startdir.append({
                'name': chunk_name,
                'size': size,
                'offset': reader.tell()
            })

            data = reader.read_string(size)
            process_table(chunk_name, data, datadir)

    # ============================================================
    # Экспорт полей (DecodeAll)
    # ============================================================

    dump = {}

    dump["fil_wrk_time"] = export_field("fil", "wrk_time", datadir, 0, 1)
    dump["wrk_time_schedule"] = export_field("wrk_time", "schedule", datadir, 1, 0)
    dump["fil_wrk_time_comment"] = export_field("fil", "wrk_time_comment", datadir, 1, 3)

    dump["rub3_rub2"] = export_field("rub3", "rub2", datadir, 0, 1)
    dump["rub2_rub1"] = export_field("rub2", "rub1", datadir, 0, 1)

    dump["rub1_name"] = export_field("rub1", "name", datadir, 1)
    dump["rub2_name"] = export_field("rub2", "name", datadir, 1)
    dump["rub3_name"] = export_field("rub3", "name", datadir, 1)

    dump["bld_purpose"] = export_field("bld_purpose", "name", datadir, 1, 0)
    dump["bld_purpose_x"] = export_field("building", "purpose", datadir, 0, 2)

    dump["bld_name"] = export_field("bld_name", "name", datadir, 1, 0)
    dump["bld_name_x"] = export_field("building", "name", datadir, 0, 2)

    dump["post_index"] = export_field("building", "post_index", datadir, 1, 3)
    dump["map_to_building"] = export_field("map_to_building", "data", datadir, 0, 4)

    dump["payment_type1"] = export_field("fil_payment", "fil", datadir, 0, 1)
    dump["payment_type2"] = export_field("fil_payment", "payment", datadir, 0, 2)
    dump["payment_type_name"] = export_field("payment_type", "name", datadir, 1)

    # Формирование типов платежей
    dump["payment"] = {}
    for key, val in dump["payment_type1"].items():
        pid = dump["payment_type2"].get(key)
        if pid is not None:
            pname = dump["payment_type_name"].get(pid, "")
            if val not in dump["payment"]:
                dump["payment"][val] = []
            dump["payment"][val].append(pname)

    dump["fil_contact_comment"] = export_field("fil_contact", "comment", datadir, 1, 3)
    dump["address_elem_map_oid"] = export_field("address_elem", "map_oid", datadir, 0, 2)

    dump["orgid"] = export_field("org", "id", datadir, 0, 2)
    dump["org"] = export_field("org", "name", datadir, 1)

    dump["orgrub_org"] = export_field("org_rub", "org", datadir, 0, 1)

    dump["fil_contact_type"] = export_field("fil_contact", "type", datadir, 0, 2)

    dump["filrub_fil"] = export_field("fil_rub", "fil", datadir, 0, 1)
    dump["filrub_rub"] = export_field("fil_rub", "rub", datadir, 0, 2)

    dump["fil_office"] = export_field("fil", "office", datadir, 1, 3)
    dump["fil_title"] = export_field("fil", "title", datadir, 1, 3)

    dump["building"] = export_field("address_elem", "building", datadir)

    dump["city"] = export_field("city", "name", datadir, 1)

    dump["orgrub_rub"] = export_field("org_rub", "rub", datadir, 0, 2)

    dump["address_elem"] = export_field("address_elem", "street", datadir, 0, 1)
    dump["street"] = export_field("street", "name", datadir, 1)
    dump["street_city"] = export_field("street", "city", datadir, 0, 1)

    dump["fil_contact_fil"] = export_field("fil_contact", "fil", datadir, 0, 1)
    dump["fil_contact_phone"] = export_field("fil_contact", "phone", datadir)
    dump["fil_contact_eaddr"] = export_field("fil_contact", "eaddr", datadir, 1)
    dump["fil_contact_eaddr_name"] = export_field("fil_contact", "eaddr_name", datadir, 1, 3)

    dump["fil_address_fil"] = export_field("fil_address", "fil", datadir, 0, 1)
    dump["fil_address_address"] = export_field("fil_address", "address", datadir, 0, 2)

    dump["fil_org"] = export_field("fil", "org", datadir, 0, 1)

    return dump, prop


# ============================================================
# Генерация XLSX из распарсенных данных
# ============================================================

def build_inverse_maps(dump: dict):
    """Строит обратные словари для оптимизации поиска."""
    fil_address_fil2 = {}
    for key, val in dump.get("fil_address_fil", {}).items():
        fil_address_fil2[val] = key

    fil_contact_fil2 = {}
    for key, val in dump.get("fil_contact_fil", {}).items():
        if val not in fil_contact_fil2:
            fil_contact_fil2[val] = []
        fil_contact_fil2[val].append(key)

    orgrub_org2 = {}
    for key, val in dump.get("orgrub_org", {}).items():
        if val not in orgrub_org2:
            orgrub_org2[val] = []
        orgrub_org2[val].append(key)

    filrub_fil2 = {}
    for key, val in dump.get("filrub_fil", {}).items():
        if val not in filrub_fil2:
            filrub_fil2[val] = []
        filrub_fil2[val].append(key)

    return fil_address_fil2, fil_contact_fil2, orgrub_org2, filrub_fil2


def parse_worktime(schedule_xml: str) -> str:
    """Парсит XML расписания работы и возвращает читаемую строку."""
    if not schedule_xml:
        return ""
    try:
        root = ET.fromstring(schedule_xml)
        lines = []
        for day in root.findall('day'):
            wh_list = day.findall('working_hours')
            if wh_list:
                label = day.get('label', '')
                parts = []
                for wh in wh_list:
                    parts.append(f"{wh.get('from', '')} - {wh.get('to', '')}")
                lines.append(f"{label}: {' '.join(parts)}")
        wrk = "\n".join(lines)
        # Перевод дней недели
        for eng, rus in [('Mon', 'Пн'), ('Tue', 'Вт'), ('Wed', 'Ср'),
                         ('Thu', 'Чт'), ('Fri', 'Пт'), ('Sat', 'Сб'), ('Sun', 'Вс')]:
            wrk = wrk.replace(eng, rus)
        return wrk.strip()
    except ET.ParseError:
        return ""


COLUMNS = [
    ("ID", 20),
    ("Название организации", 40),
    ("Населенный пункт", 20),
    ("Раздел", 40),
    ("Подраздел", 40),
    ("Рубрика", 40),
    ("Телефоны", 30),
    ("Факсы", 30),
    ("Email", 20),
    ("Сайт", 20),
    ("Адрес", 30),
    ("Почтовый индекс", 10),
    ("Типы платежей", 20),
    ("Время работы", 34),
    ("Собственное название строения", 25),
    ("Назначение строения", 25),
    ("Vkontakte", 20),
    ("Facebook", 20),
    ("Skype", 20),
    ("Twitter", 20),
    ("Instagram", 20),
    ("ICQ", 20),
    ("Jabber", 20),
]


def build_all_rows(dump: dict, default_city: str = "") -> list:
    """Формирует список строк (list of lists) из распарсенных данных dgdat.
    Каждая строка — список из 23 значений, соответствующих COLUMNS.

    default_city — название города из заголовка dgdat (fallback, если город
    не удалось определить через цепочку адрес → улица → город)."""

    fil_address_fil2, fil_contact_fil2, orgrub_org2, filrub_fil2 = build_inverse_maps(dump)

    fil_org = dump.get("fil_org", {})
    rows = []
    prev_id = None
    prev_data = {}

    for key, fil in sorted(fil_org.items()):
        payments_list = dump.get("payment", {}).get(key, [])
        payments = "\n".join(payments_list) if payments_list else ""

        name = dump.get("org", {}).get(fil, "")
        org_id = dump.get("orgid", {}).get(fil, "")

        # Адрес
        addr_row = fil_address_fil2.get(key)
        addr_elem_row = dump.get("fil_address_address", {}).get(addr_row) if addr_row is not None else None

        building = dump.get("building", {}).get(addr_elem_row, "") if addr_elem_row else ""
        map_oid = dump.get("address_elem_map_oid", {}).get(addr_elem_row) if addr_elem_row else None
        map_to_bld = dump.get("map_to_building", {}).get(map_oid) if map_oid else None
        post_index = dump.get("post_index", {}).get(map_to_bld, "") if map_to_bld else ""

        street_row = dump.get("address_elem", {}).get(addr_elem_row) if addr_elem_row else None
        street_name = dump.get("street", {}).get(street_row, "") if street_row else ""
        street_city_id = dump.get("street_city", {}).get(street_row) if street_row else None
        cityname = dump.get("city", {}).get(street_city_id, "") if street_city_id else ""

        # Fallback: если город не определён через цепочку адресов,
        # используем название города из заголовка dgdat-файла
        if not cityname and default_city:
            cityname = default_city

        # Контакты
        phones = []
        faxes = []
        wwws = []
        emails = []
        links = {}

        contact_rows = fil_contact_fil2.get(key, [])
        for crow in contact_rows:
            ctype_val = dump.get("fil_contact_type", {}).get(crow)
            if ctype_val is None:
                continue
            ctype = chr(ctype_val)

            if ctype == 'p':
                phone = dump.get("fil_contact_phone", {}).get(crow, "")
                if phone:
                    phones.append(phone)
            if ctype == 'f':
                phone = dump.get("fil_contact_phone", {}).get(crow, "")
                if phone:
                    faxes.append(phone)

            www = dump.get("fil_contact_eaddr_name", {}).get(crow, "")
            if www:
                wwws.append(www.lower() if isinstance(www, str) else www)

            eaddr = dump.get("fil_contact_eaddr", {}).get(crow, "")
            if isinstance(eaddr, str):
                eaddr = eaddr.lower()

            if ctype == 'm' and eaddr:
                emails.append(eaddr)

            if ctype in ('t', 'v', 'a', 'n', 's', 'i', 'j'):
                raw_eaddr = dump.get("fil_contact_eaddr", {}).get(crow, "")
                if raw_eaddr:
                    if ctype not in links:
                        links[ctype] = []
                    links[ctype].append(raw_eaddr)

        # Рубрики
        rubs3 = []
        rubs2 = []
        rubs1 = []

        org_rub_rows = orgrub_org2.get(fil, [])
        fil_rub_rows = filrub_fil2.get(key, [])

        for r in org_rub_rows:
            rubid = dump.get("orgrub_rub", {}).get(r)
            if rubid is not None:
                rubs3.append(dump.get("rub3_name", {}).get(rubid, ""))
                rub2id = dump.get("rub3_rub2", {}).get(rubid)
                if rub2id is not None:
                    rubs2.append(dump.get("rub2_name", {}).get(rub2id, ""))
                    rub1id = dump.get("rub2_rub1", {}).get(rub2id)
                    if rub1id is not None:
                        rubs1.append(dump.get("rub1_name", {}).get(rub1id, ""))

        for r in fil_rub_rows:
            rubid = dump.get("filrub_rub", {}).get(r)
            if rubid is not None:
                rubs3.append(dump.get("rub3_name", {}).get(rubid, ""))
                rub2id = dump.get("rub3_rub2", {}).get(rubid)
                if rub2id is not None:
                    rubs2.append(dump.get("rub2_name", {}).get(rub2id, ""))
                    rub1id = dump.get("rub2_rub1", {}).get(rub2id)
                    if rub1id is not None:
                        rubs1.append(dump.get("rub1_name", {}).get(rub1id, ""))

        rubs3_str = "\n".join(sorted(set(rubs3)))
        rubs2_str = "\n".join(sorted(set(rubs2)))
        rubs1_str = "\n".join(sorted(set(rubs1)))

        phones_str = "\n".join(phones)
        faxes_str = "\n".join(faxes)
        wwws_str = "\n".join(wwws)
        emails_str = "\n".join(emails)
        vk = "\n".join(links.get('v', []))
        fb = "\n".join(links.get('a', []))
        skype = "\n".join(links.get('s', []))
        twitter = "\n".join(links.get('t', []))
        insta = "\n".join(links.get('n', []))
        icq = "\n".join(links.get('i', []))
        jabber = "\n".join(links.get('j', []))

        # Время работы
        wt_id = dump.get("fil_wrk_time", {}).get(key)
        wt_schedule = dump.get("wrk_time_schedule", {}).get(wt_id, "") if wt_id else ""
        wrk = parse_worktime(wt_schedule)

        # Наследование от предыдущей записи с тем же ID
        if prev_id == org_id:
            if not emails_str:
                emails_str = prev_data.get('emails', '')
            if not wwws_str:
                wwws_str = prev_data.get('wwws', '')
            if not vk:
                vk = prev_data.get('vk', '')
            if not twitter:
                twitter = prev_data.get('twitter', '')
            if not fb:
                fb = prev_data.get('fb', '')
            if not insta:
                insta = prev_data.get('insta', '')
            if not skype:
                skype = prev_data.get('skype', '')
            if not icq:
                icq = prev_data.get('icq', '')

        address = street_name
        if building:
            address = f"{street_name}, {building}" if street_name else building

        if name and isinstance(name, str) and name.startswith("="):
            name = name[1:]

        bld_purpose_id = dump.get("bld_purpose_x", {}).get(key)
        bld_purpose = dump.get("bld_purpose", {}).get(bld_purpose_id, "") if bld_purpose_id else ""
        bld_name_id = dump.get("bld_name_x", {}).get(key)
        bld_name = dump.get("bld_name", {}).get(bld_name_id, "") if bld_name_id else ""

        values = [
            org_id, name, cityname, rubs1_str, rubs2_str, rubs3_str,
            phones_str, faxes_str, emails_str, wwws_str, address,
            post_index, payments, wrk, bld_name, bld_purpose,
            vk, fb, skype, twitter, insta, icq, jabber
        ]

        # Очистка управляющих символов
        values = [ILLEGAL_CHARS_RE.sub('', v) if isinstance(v, str) else v for v in values]

        rows.append(values)

        prev_id = org_id
        prev_data = {
            'emails': emails_str, 'wwws': wwws_str, 'vk': vk,
            'twitter': twitter, 'fb': fb, 'insta': insta,
            'skype': skype, 'icq': icq, 'jabber': jabber,
        }

    return rows


def _normalize_cell(val):
    """Нормализует значение ячейки для сравнения."""
    if val is None:
        return ''
    if isinstance(val, (int, float)):
        return val
    return str(val)


def _setup_sheet_headers(ws):
    """Настраивает заголовки и стили листа."""
    header_fill = PatternFill(start_color="FFC4D79B", end_color="FFC4D79B", fill_type="solid")
    header_font = Font(name='Arial', size=10, color="0000FF")
    header_align = Alignment(vertical='center', wrap_text=True, horizontal='left')

    for col_idx, (col_name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 50
    ws.freeze_panes = 'A2'
    last_col_letter = ws.cell(row=1, column=len(COLUMNS)).column_letter
    ws.auto_filter.ref = f'A1:{last_col_letter}1'


def write_xlsx(rows: list, output_path: str):
    """Создаёт XLSX файл из списка строк."""
    wb = Workbook()
    ws = wb.active
    _setup_sheet_headers(ws)

    default_font = Font(name='Arial', size=10)
    default_align = Alignment(wrap_text=True, vertical='top', horizontal='left')

    for row_idx, values in enumerate(rows, 2):
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = default_font
            cell.alignment = default_align

        if row_idx % 1000 == 0:
            print(f"  {row_idx - 1}/{len(rows)}")

    wb.save(output_path)


def xlsx_matches_rows(rows: list, xlsx_path: str) -> bool:
    """Сравнивает данные в существующем XLSX с новыми строками.
    Возвращает True, если данные полностью совпадают."""
    wb = load_workbook(xlsx_path)
    ws = wb.active

    num_cols = len(COLUMNS)
    existing_row_count = ws.max_row - 1  # минус заголовок

    if existing_row_count != len(rows):
        return False

    for row_idx, new_values in enumerate(rows, 2):
        for col_idx, new_val in enumerate(new_values, 1):
            old_val = _normalize_cell(ws.cell(row=row_idx, column=col_idx).value)
            new_val_norm = _normalize_cell(new_val)
            if old_val != new_val_norm:
                return False

    return True


# ============================================================
# CLI
# ============================================================

def main():
    os.makedirs(XLSX_DIR, exist_ok=True)

    if len(sys.argv) >= 2:
        # Явно указан файл
        input_file = sys.argv[1]
        if not os.path.exists(input_file):
            print(f"Файл не найден: {input_file}")
            sys.exit(1)

        if len(sys.argv) >= 3:
            output_file = sys.argv[2]
        else:
            base = os.path.splitext(os.path.basename(input_file))[0]
            name_part = base.split('-')[0] if '-' in base else base
            output_file = os.path.join(XLSX_DIR, name_part + ".xlsx")

        convert_file(input_file, output_file)
    else:
        # Без аргументов — обработать все .dgdat из DGDAT_DIR
        if not os.path.isdir(DGDAT_DIR):
            print(f"Папка {DGDAT_DIR} не найдена.")
            print()
            print("Использование:")
            print(f"  python {os.path.basename(__file__)}                          — конвертирует все .dgdat из DGDAT_DIR")
            print(f"  python {os.path.basename(__file__)} <файл.dgdat>             — конвертирует один файл")
            print(f"  python {os.path.basename(__file__)} <файл.dgdat> <выход.xlsx> — с указанием выходного файла")
            sys.exit(1)

        dgdat_files = sorted(
            f for f in os.listdir(DGDAT_DIR) if f.lower().endswith('.dgdat')
        )

        if not dgdat_files:
            print(f"Нет .dgdat файлов в {DGDAT_DIR}")
            sys.exit(0)

        print(f"Найдено файлов: {len(dgdat_files)}")
        print(f"  Вход: {DGDAT_DIR}")
        print(f"  Выход: {XLSX_DIR}")

        for dgdat in dgdat_files:
            input_path = os.path.join(DGDAT_DIR, dgdat)
            name_part = dgdat.split('-')[0] if '-' in dgdat else os.path.splitext(dgdat)[0]
            output_path = os.path.join(XLSX_DIR, name_part + ".xlsx")

            convert_file(input_path, output_path)

        print("\nВсе файлы обработаны.")


def convert_file(input_file: str, output_file: str):
    """Конвертирует один .dgdat файл в .xlsx.
    Если xlsx уже существует и данные совпадают — пропускает.
    Если есть расхождения или файла нет — создаёт/перезаписывает."""
    print(f"\nПарсинг: {input_file}")
    dump, prop = parse_dgdat(input_file)
    print(f"  Город: {prop.get('name', '?')}")
    print("  Формирование данных...")
    rows = build_all_rows(dump, default_city=prop.get("name", ""))
    print(f"  Записей в dgdat: {len(rows)}")

    if os.path.exists(output_file):
        print(f"  Сравнение с {os.path.basename(output_file)}...")
        if xlsx_matches_rows(rows, output_file):
            print("  Данные актуальны, обновление не требуется.")
            return
        print("  Обнаружены расхождения — перезапись...")

    write_xlsx(rows, output_file)
    print(f"  Сохранено: {output_file} ({len(rows)} записей)")


if __name__ == '__main__':
    main()
